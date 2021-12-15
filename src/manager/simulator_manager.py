import multiprocessing, time, traceback, datetime
from src.define import COLLECT_MODULES
from collections import defaultdict
from src.filter import Filter
from src.detector import Detector
from .simulator_shared_data import SimulatorSharedData
from src.manager.context import Context
from src.detector.detect_list import DETECT_LIST
from src.util.db.soul import SoulDBCon
from src.util.excel_module import ExcelModule
from openpyxl.styles import Alignment, Border, borders, PatternFill, Color
from openpyxl.styles.colors import BLUE

class SimulatorManager:
    def __init__(self):
        self._logger = multiprocessing.get_logger()
        self._modules = COLLECT_MODULES

    def make_coll_names(cls, col_name, date_at_from, date_at_to):
        s_y = int(date_at_from[:2])
        s_m = int(date_at_from[2:4])
        e_y = int(date_at_to[:2])
        e_m = int(date_at_to[2:4])
        coll_names = []
        while True:
            if (s_y > e_y) or (s_y == e_y and s_m > e_m):
                break
            str_y = f'{"0" if s_y < 10 else ""}{str(s_y)}'
            str_m = f'{"0" if s_m < 10 else ""}{str(s_m)}'
            coll_names.append(col_name + str_y + str_m)
            s_m += 1
            if s_m > 12:
                s_y += 1
                s_m = 1
        return coll_names

    def find(self, db_obj, col_name, mcno, datetime_from, datetime_to):
        col_names = self.make_coll_names(col_name, datetime_from, datetime_to)
        logs = []
        query = {
            'mcno': mcno,
            'datetimeAt': {'$gte': datetime_from, '$lte': datetime_to}
        }
        for col_name in col_names:
            col_obj = db_obj.get_collection(col_name)
            docs = list(col_obj.find(query).sort('datetimeAt', 1))
            logs.extend(docs)
        return logs

    def load_data(self, mcno, datetime_from, datetime_to):
        data_map = defaultdict(dict)
        datetime_ats = []
        log_cnt = 0

        for module in self._modules:
            db_obj = SoulDBCon.getHotDB(module)
            logs = self.find(db_obj, f'{module}Log_', mcno, datetime_from, datetime_to)
            log_cnt += len(logs)
            for log in logs:
                datetime_at = log.get('datetimeAt')
                if datetime_at not in datetime_ats:
                    datetime_ats.append(datetime_at)
                data_map[datetime_at][module] = log
        datetime_ats.sort(key=lambda datetime_at: int(datetime_at))

        return datetime_ats, data_map, log_cnt

    def datetime_at_to_datetime(self, datetime_at):
        return datetime.datetime(
            int('21' + datetime_at[:2]),
            int(datetime_at[2:4]),
            int(datetime_at[4:6]),
            int(datetime_at[6:8]),
            int(datetime_at[8:10]),
            int(datetime_at[10:12]),
        )

    def create_context(self, log, module, shared_data, base_time):
        context = Context()
        context.raw = log
        context.timestamp = base_time
        context.messageType = module
        return context

    def check_close(self, shared_data, code_to_detector, base_time):
        events = shared_data.getCategoryObj('event')
        close_events = []

        if events:
            for key in events:
                event = events.get(key)
                mcno = event.get('mcno')
                module = event.get('module')
                eventCode = event.get('eventCode')

                data = shared_data.get(module, mcno)

                if not data:
                    continue

                detector = code_to_detector.get(eventCode)

                if not detector:
                    continue

                timestamp = base_time
                detectEvent = detector.detect(data, timestamp)

                if not detectEvent:
                    event['endTime'] = base_time
                    event['state'] = 'close'
                    close_events.append(event)

        return close_events

    def check_comm_error(self, shared_data, comm_error_detectors, base_time):
        context_list = []
        for module in self._modules:
            docs = shared_data.getCategoryObj(module)

            if not docs:
                continue

            detector = comm_error_detectors.get(module)

            if not detector:
                continue

            for mcno in docs:
                doc = docs.get(mcno)
                timestamp = base_time
                event = detector.detect(doc, timestamp)

                if not event:
                    continue

                context = self.create_context(doc, module, shared_data, base_time)
                context.messageType = 'event'
                context.openEvents.append(event)
                context_list.append(context)
        return context_list

    def run(self, mcno, datetime_from, datetime_to):
        s = time.time()
        try:
            print('Init shared data...')
            shared_data = SimulatorSharedData()
            shared_data.init(mcno, shared_data.getDataObj())
            shared_data_init_time = time.time() - s
            print('Complete init shared data! time: {} sec'.format(shared_data_init_time))

            detector_object = Detector()
            code_to_detector = detector_object.getCodeToDetector(shared_data)
            filter_object = Filter()
            detectors = DETECT_LIST(shared_data)
            comm_error_detectors = detector_object.getCommErorDetectors(shared_data)
            print('Load Logs...')
            s = time.time()
            datetime_ats, data_map, log_cnt = self.load_data(mcno, datetime_from, datetime_to)
            load_log_time = time.time() - s
            print('Complete Load Logs! time: {} sec'.format(load_log_time))

            print(f'{log_cnt} logs')

            if log_cnt == 0:
                return 'log가 존재하지 않습니다.'

            open_events = []
            close_events = []
            error_list = []

            s = time.time()
            print('Processing...')
            for datetime_at in datetime_ats:
                base_time = self.datetime_at_to_datetime(datetime_at)
                context_list = []
                for module in self._modules:
                    if not data_map.get(datetime_at):
                        continue
                    if not data_map.get(datetime_at).get(module):
                        continue
                    log = data_map.get(datetime_at).get(module)

                    if log:
                        context = self.create_context(log, module, shared_data, base_time)
                        detect_result = detector_object.process(context, detectors)
                        shared_data.update(context.messageType, context.raw.get('mcno'), context.raw)
                        if detect_result == 'true':
                            context_list.append(context)
                        elif detect_result not in ['true', 'false']:
                            error_list.append({'context': context, 'error': detect_result})

                detect_close_events = self.check_close(shared_data, code_to_detector, base_time)

                for context in context_list:
                    filter_object.process(context, shared_data)
                    for open_event in context.openEvents:
                        key = shared_data.eventKey(open_event.get('mcno'), open_event.get('eventCode'))
                        shared_data.put('event', key, open_event)
                    open_events.extend(context.openEvents)

                for close_event in detect_close_events:
                    key = shared_data.eventKey(close_event.get('mcno'), close_event.get('eventCode'))
                    shared_data.delete('event', key)
                    close_events.append(close_event)

            process_time = time.time() - s
            print('Complete processing! time: {} sec'.format(process_time))
            close_event_ids = list(map(lambda el: el.get('eventId'), close_events))
            open_events = list(filter(lambda el: (el.get('eventId') not in close_event_ids), open_events))
            total_events = open_events + close_events
            total_events.sort(key=lambda el: el.get('startTime'), reverse=True)

            print('----------[Summary]----------')
            print('Init shared data time: {}'.format(shared_data_init_time))
            print('Load log time: {}'.format(load_log_time))
            print('Process time: {}'.format(process_time))
            print('Detect anomaly: {}'.format(len(open_events) + len(close_events)))
            print('Close anomaly: {}'.format(len(close_events)))
            print('Open anomaly: {}'.format(len(open_events)))

            print('----------[OPEN EVENT]----------')
            for open_event in open_events:
                print(open_event)

            print('----------[CLOSE EVENT]----------')
            for close_event in close_events:
                print(close_event)

            print('----------[ERROR]----------')
            for error in error_list:
                print(error)

            self.export_excel(
                mcno,
                datetime_from,
                datetime_to,
                shared_data_init_time,
                load_log_time,
                process_time,
                total_events,
                open_events,
                close_events,
                error_list
            )

        except Exception as e:
            print(traceback.format_exc())

    def export_excel(self,
                     mcno,
                     datetime_from,
                     datetime_to,
                     shared_data_init_time,
                     load_log_time,
                     process_time,
                     total_events,
                     open_events,
                     close_events,
                     error_list):
        excel_module = ExcelModule(f'{mcno}_{datetime_from}_{datetime_to}')
        ws = excel_module.getCurWorkSheet()
        self.excel_summary(
            ws,
            mcno,
            datetime_from,
            datetime_to,
            shared_data_init_time,
            load_log_time,
            process_time,
            open_events,
            close_events,
            error_list
        )
        self.excel_log(excel_module, total_events)
        excel_module.save()

    def excel_summary(self,
                      ws,
                      mcno,
                      datetime_from,
                      datetime_to,
                      shared_data_init_time,
                      load_log_time,
                      process_time,
                      open_events,
                      close_events,
                      error_list):

        db_obj = SoulDBCon.getHotDB('meta')
        ps_obj = db_obj.get_collection('powerStations').find_one({'mcno': mcno}, {'psName': True})
        ps_name = '{}({})'.format(ps_obj.get('psName'), mcno)

        ws.title = 'Summary'
        ws.merge_cells('B2:C2')
        ws.merge_cells('B3:C3')
        ws.merge_cells('B4:B6')
        ws.merge_cells('B7:B9')
        ws.merge_cells('B10:C10')
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 24

        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['H'].width = 100

        THIN_BORDER = Border(borders.Side('thin'), borders.Side('thin'), borders.Side('thin'), borders.Side('thin'))
        CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

        ws['B2'].value = '대상 발전소'
        ws['B2'].alignment = CENTER_ALIGNMENT
        ws['B2'].border = THIN_BORDER
        ws['C2'].border = THIN_BORDER
        ws['D2'].value = ps_name
        ws['D2'].alignment = CENTER_ALIGNMENT
        ws['D2'].border = THIN_BORDER
        ws['B3'].value = '대상 구간'
        ws['B3'].alignment = CENTER_ALIGNMENT
        ws['B3'].border = THIN_BORDER
        ws['D3'].value = '{}\n ~{}'.format(
            self.datetime_at_to_datetime(datetime_from).strftime('%Y/%m/%d %H:%M'),
            self.datetime_at_to_datetime(datetime_to).strftime('%Y/%m/%d %H:%M'))
        ws['D3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['D3'].border = THIN_BORDER
        ws['B4'].value = '수행 시간'
        ws['B4'].alignment = CENTER_ALIGNMENT
        ws['B4'].border = THIN_BORDER
        ws['B5'].border = THIN_BORDER
        ws['B6'].border = THIN_BORDER
        ws['C4'].value = '초기 데이터 로드'
        ws['C4'].alignment = CENTER_ALIGNMENT
        ws['C4'].border = THIN_BORDER
        ws['D4'].value = shared_data_init_time
        ws['D4'].alignment = CENTER_ALIGNMENT
        ws['D4'].border = THIN_BORDER
        ws['C5'].value = '로그 데이터 로드'
        ws['C5'].alignment = CENTER_ALIGNMENT
        ws['C5'].border = THIN_BORDER
        ws['D5'].value = load_log_time
        ws['D5'].alignment = CENTER_ALIGNMENT
        ws['D5'].border = THIN_BORDER
        ws['C6'].value = '이벤트 처리'
        ws['C6'].alignment = CENTER_ALIGNMENT
        ws['C6'].border = THIN_BORDER
        ws['D6'].value = process_time
        ws['D6'].alignment = CENTER_ALIGNMENT
        ws['D6'].border = THIN_BORDER
        ws['B7'].value = '이벤트'
        ws['B7'].alignment = CENTER_ALIGNMENT
        ws['B7'].border = THIN_BORDER
        ws['B8'].border = THIN_BORDER
        ws['B9'].border = THIN_BORDER
        ws['C7'].value = 'TOTAL'
        ws['C7'].alignment = CENTER_ALIGNMENT
        ws['C7'].border = THIN_BORDER
        ws['D7'].value = len(open_events) + len(close_events)
        ws['D7'].alignment = CENTER_ALIGNMENT
        ws['D7'].border = THIN_BORDER
        ws['C8'].value = 'OPEN'
        ws['C8'].alignment = CENTER_ALIGNMENT
        ws['C8'].border = THIN_BORDER
        ws['D8'].value = len(open_events)
        ws['D8'].alignment = CENTER_ALIGNMENT
        ws['D8'].border = THIN_BORDER
        ws['C9'].value = 'CLOSE'
        ws['C9'].alignment = CENTER_ALIGNMENT
        ws['C9'].border = THIN_BORDER
        ws['D9'].value = len(close_events)
        ws['D9'].alignment = CENTER_ALIGNMENT
        ws['D9'].border = THIN_BORDER
        ws['B10'].value = '에러건'
        ws['B10'].alignment = CENTER_ALIGNMENT
        ws['B10'].border = THIN_BORDER
        ws['C10'].border = THIN_BORDER
        ws['D10'].value = len(error_list)
        ws['D10'].alignment = CENTER_ALIGNMENT
        ws['D10'].border = THIN_BORDER

        event_codes, event_summary = self.process_event_summary(open_events, close_events)
        start_row = 4
        start_column = 6

        ws['F3'].value = '이벤트명'
        ws['F3'].alignment = CENTER_ALIGNMENT
        ws['F3'].border = THIN_BORDER
        ws['G3'].value = '코드'
        ws['G3'].alignment = CENTER_ALIGNMENT
        ws['G3'].border = THIN_BORDER
        ws['H3'].value = '설명'
        ws['H3'].alignment = CENTER_ALIGNMENT
        ws['H3'].border = THIN_BORDER
        ws['I3'].value = 'TOTAL'
        ws['I3'].alignment = CENTER_ALIGNMENT
        ws['I3'].border = THIN_BORDER
        ws['J3'].value = 'OPEN'
        ws['J3'].alignment = CENTER_ALIGNMENT
        ws['J3'].border = THIN_BORDER
        ws['K3'].value = 'CLOSE'
        ws['K3'].alignment = CENTER_ALIGNMENT
        ws['K3'].border = THIN_BORDER

        for idx, event_code in enumerate(event_codes):
            event_obj = event_summary[event_code]
            for idx2, key in enumerate(['eventName', 'eventCode', 'eventDesc', 'totalEvents', 'openEvents', 'closeEvents']):
                cell = ws.cell(row=start_row + idx, column=start_column + idx2, value = event_obj.get(key))
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT
                if key in ['totalEvents', 'openEvents', 'closeEvents']:
                    if event_obj.get(key) > 0:
                        cell.fill = PatternFill(start_color='faceac', end_color='faceac', fill_type='solid')

    def process_event_summary(self, open_events, close_events):
        db_obj = SoulDBCon.getHotDB('meta')
        event_specs = list(db_obj.get_collection('eventSpec').find({'eventCode': {'$gte': 2000, '$lt': 6000}}))
        event_codes = list(map(lambda el: el.get('eventCode'), event_specs))
        event_codes.sort()
        event_map = {}
        for event_spec in event_specs:
            event_map[event_spec.get('eventCode')] = {
                'eventCode': event_spec.get('eventCode'),
                'eventName': event_spec.get('eventName'),
                'eventDesc': event_spec.get('eventDesc'),
                'totalEvents': 0,
                'openEvents': 0,
                'closeEvents': 0
            }
        for event in open_events:
            event_map[event.get('eventCode')]['totalEvents'] += 1
            event_map[event.get('eventCode')]['openEvents'] += 1
        for event in close_events:
            event_map[event.get('eventCode')]['totalEvents'] += 1
            event_map[event.get('eventCode')]['closeEvents'] += 1
        return event_codes, event_map

    def excel_log(self, excel_module, total_events):
        excel_module.createSheet('log')
        ws = excel_module.getCurWorkSheet()
        THIN_BORDER = Border(borders.Side('thin'), borders.Side('thin'), borders.Side('thin'), borders.Side('thin'))
        CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['F'].width = 24
        ws.column_dimensions['G'].width = 24
        ws.column_dimensions['H'].width = 100

        for idx, title in enumerate(['이벤트명', '코드', '모듈', '모듈NO', '발생시간', '종료시간', '메시지']):
            cell = ws.cell(row=2, column=2 + idx, value=title)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT

        for row_idx, event in enumerate(total_events):
            for col_idx, key in enumerate(['eventName', 'eventCode', 'module', 'moduleNos', 'startTime', 'endTime', 'message']):
                cell = ws.cell(row=3 + row_idx, column=2 + col_idx, value=str(event.get(key)))
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGNMENT