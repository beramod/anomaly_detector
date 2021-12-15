import openpyxl

class ExcelModule:
    def __init__(self, fileName):
        self._fileName = fileName
        self._wb = openpyxl.Workbook()
        self._ws = self._wb.active
        self._offset = 1

    def getCurWorkSheet(self):
        return self._ws

    def setLineAlignment(self):
        self._offset += 1

    def setTitles(self, titles: list):
        for idx, title in enumerate(titles):
            self._ws.cell(row=self._offset, column=idx + 2, value=title)

    def createSheet(self, sheetName):
        self._ws = self._wb.create_sheet(sheetName)

    def selectSheet(self, sheetName):
        self._ws = self._wb[sheetName]

    def setContents(self, contents: list):
        for idx1, content in enumerate(contents):
            for idx2, each in enumerate(content):
                self._ws.cell(row=idx1 + 2, column=idx2 + 2, value=each)

    def save(self):
        self._wb.save('files/{}.xlsx'.format(self._fileName))